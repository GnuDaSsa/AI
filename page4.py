import streamlit as st
import openpyxl
import pandas as pd
from datetime import datetime
import win32com.client
import pythoncom
import io
import zipfile
import tempfile
import os

# 이 파일의 모든 Streamlit UI 및 로직은 run() 함수 내에 정의됩니다.
def run():
    # COM 라이브러리 초기화는 run() 함수 시작 시 수행
    # HWP 관련 작업이 이 함수 내에서만 이루어지므로 여기에 두는 것이 적절합니다.
    pythoncom.CoInitialize()

    # 페이지 설정 및 제목은 메인 앱에서 이미 처리되므로 여기서는 제거합니다.
    # st.set_page_config(layout="wide")
    # st.title("도급위탁용역 점검표 생성기")

    # 템플릿 파일 경로 정의 (시트별로 다르게 사용)
    template_paths_by_sheet = {
        "점검표1": r"C:\\Users\\Owner\\Desktop\\사진우\\AI\\도급위탁용역자동화\\서식\\점검표(60일 이상).hwp",
        "점검표2": r"C:\\Users\Owner\\Desktop\\사진우\\AI\\도급위탁용역자동화\\서식\\점검표(60일 이하).hwp",
        "점검표3": r"C:\\Users\\Owner\\Desktop\\사진우\\AI\\도급위탁용역자동화\\서식\\점검표3.hwp"
    }

    # 의무(tag) 그룹 전역 정의 (전체 가능한 그룹)
    full_compliance_groups = [
        (17, 18, 19, ["[Q]", "[R]", "[S]"]),
        (20, 21, 22, ["[T]", "[U]", "[V]"]),
        (23, 24, 25, ["[W]", "[X]", "[Y]"]),
        (26, 27, 28, ["[Z]", "[AA]", "[AB]"]),
        (29, 30, 31, ["[AC]", "[AD]", "[AE]"]),
        (32, 33, 34, ["[AF]", "[AG]", "[AH]"]),
        (35, 36, 37, ["[AI]", "[AJ]", "[AK]"]), # Group 7 (index 6)
        (38, 39, 40, ["[AL]", "[AM]", "[AN]"]), # Group 8 (index 7)
        (41, 42, 43, ["[AO]", "[AP]", "[AQ]"]), # Group 9 (index 8) - Assuming these columns/tags
        (44, 45, 46, ["[AR]", "[AS]", "[AT]"])  # Group 10 (index 9) - Assuming these columns/tags
    ]

    # 시트별 의무(tag) 그룹 정의
    compliance_groups_by_sheet = {
        "점검표1": full_compliance_groups, # 점검표1은 전체 그룹 사용
        "점검표2": full_compliance_groups[:7], # 점검표2는 AI, AJ, AK 그룹까지만 사용 (인덱스 0-6)
        "점검표3": full_compliance_groups # 점검표3은 AR, AS, AT 그룹까지 사용 (전체 그룹)
    }

    # CSS: 입력 상자 최대 너비 및 placeholder 스타일
    st.markdown("""
    <style>
    div.stTextInput > div > input {
        max-width: 200px;
    }
    [data-baseweb="input"] > input::placeholder {
        color: #999 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # 세션 상태 초기화 (각 페이지의 세션 상태 키는 고유해야 합니다.)
    if "extracted_page4" not in st.session_state:
        st.session_state.extracted_page4 = [] # 모든 시트에서 추출된 데이터를 담을 리스트
    if "dates_page4" not in st.session_state:
        st.session_state.dates_page4 = []
    if "apply_all_checkbox_page4" not in st.session_state:
        st.session_state.apply_all_checkbox_page4 = False
    if "extracted_data_by_sheet_page4" not in st.session_state:
        st.session_state.extracted_data_by_sheet_page4 = {} # 시트별로 구분된 데이터를 담을 딕셔너리

    left_col, right_col = st.columns([1, 1])

    with left_col:
        # 팀장/과장 입력란을 한 줄에 배치 (key 변경)
        c1, c2 = st.columns([1, 1])
        with c1:
            team_leader = st.text_input("팀장님 성함", max_chars=10, key="team_leader_page4")
        with c2:
            manager     = st.text_input("과장님 성함", max_chars=10, key="manager_page4")

        uploaded_file = st.file_uploader("엑셀 파일을 업로드하세요", type=["xlsx"], key="uploaded_file_page4") # key 변경
        generate_btn = st.button("생성", key="generate_btn_page4") # key 변경

        # 입력 검증
        if generate_btn and (not team_leader or not manager):
            st.warning("팀장님과 과장님 성함을 모두 입력하세요.")
        elif generate_btn and not uploaded_file:
            st.warning("엑셀 파일을 업로드하세요.")

        # 데이터 추출 및 화면 표시
        if generate_btn and team_leader and manager and uploaded_file:
            # 데이터 추출 함수 정의
            def extract_data_from_sheets():
                wb = openpyxl.load_workbook(uploaded_file)
                # 점검표1, 점검표2, 점검표3 시트에서 데이터를 추출하도록 변경
                target_sheet_names_prefixes = ["점검표1", "점검표2", "점검표3"]
                all_extracted_data_flat = [] # 모든 시트에서 추출된 데이터를 하나의 리스트로
                extracted_data_by_sheet = {} # 시트별로 구분된 데이터를 딕셔너리로

                for sheet_prefix in target_sheet_names_prefixes:
                    # 실제 워크북에 존재하는 시트 이름 찾기 (부분 일치)
                    actual_sheet_name = next((s for s in wb.sheetnames if sheet_prefix in s), None)

                    if not actual_sheet_name:
                        # 해당 시트가 없으면 경고를 표시하고 다음 시트로 넘어갑니다.
                        st.warning(f"'{sheet_prefix}' 시트를 찾을 수 없습니다. 다음 시트로 넘어갑니다.")
                        continue

                    sh = wb[actual_sheet_name]
                    current_sheet_extracted_data = [] # 현재 시트에서 추출된 데이터

                    # 현재 시트에 해당하는 compliance_groups를 가져옵니다.
                    sheet_compliance_groups = compliance_groups_by_sheet.get(sheet_prefix, full_compliance_groups)

                    for row in sh.iter_rows(min_row=2, max_row=sh.max_row, min_col=2, max_col=2):
                        cell = row[0]
                        # 셀 값이 존재하고, 폰트 및 색상이 빨간색인 경우만 추출
                        if cell.value and cell.font and cell.font.color and cell.font.color.rgb == "FFFF0000":
                            # 기존 포맷팅 로직 유지
                            mgr_raw   = sh.cell(row=cell.row, column=6).value or ""
                            start_raw = sh.cell(row=cell.row, column=8).value
                            end_raw   = sh.cell(row=cell.row, column=9).value
                            cost_raw  = sh.cell(row=cell.row, column=11).value

                            def fmt_date(d):
                                if isinstance(d, datetime):
                                    return d.strftime("%Y.%m.%d.")
                                return d

                            start = fmt_date(start_raw)
                            end   = fmt_date(end_raw)
                            cost  = "{:,}".format(int(cost_raw)) if isinstance(cost_raw, (int, float)) else cost_raw
                            mgr_name = mgr_raw.split('(')[0].strip()

                            row_data = {
                                "사업명": sh.cell(row=cell.row, column=2).value,
                                "부서명": sh.cell(row=cell.row, column=5).value,
                                "담당자": mgr_name,
                                "착공일": start,
                                "준공일": end,
                                "사업비": cost,
                                "팀장님": team_leader,
                                "과장님": manager,
                                "source_sheet": sheet_prefix # 어떤 시트에서 추출되었는지 정보 추가
                            }

                            # M1/M2 처리
                            m = sh.cell(row=cell.row, column=13).value
                            if str(m).strip().upper() == "O":
                                row_data["M1"], row_data["M2"] = "○", ""
                            elif str(m).strip().upper() == "X":
                                row_data["M1"], row_data["M2"] = "", "○"
                            else:
                                row_data["M1"], row_data["M2"] = "", ""

                            # 의무 그룹 처리 (현재 시트에 해당하는 그룹만 사용)
                            for i, (c1_, c2_, c3_, tags) in enumerate(sheet_compliance_groups, start=1):
                                v1 = sh.cell(row=cell.row, column=c1_).value
                                v2 = sh.cell(row=cell.row, column=c2_).value
                                v3 = sh.cell(row=cell.row, column=c3_).value
                                if v1: row_data[f"의무{i}"] = tags[0]
                                elif v2: row_data[f"의무{i}"] = tags[1]
                                elif v3: row_data[f"의무{i}"] = tags[2]
                                else:    row_data[f"의무{i}"] = None

                            current_sheet_extracted_data.append(row_data)
                            all_extracted_data_flat.append(row_data) # 전체 리스트에도 추가
                    
                    if current_sheet_extracted_data:
                        extracted_data_by_sheet[sheet_prefix] = current_sheet_extracted_data

                return all_extracted_data_flat, extracted_data_by_sheet

            st.session_state.extracted_page4, st.session_state.extracted_data_by_sheet_page4 = extract_data_from_sheets() # 세션 상태 키 변경
            
            if not st.session_state.extracted_page4:
                st.warning("조건을 만족하는 데이터가 없습니다.")
            else:
                # 시트별 추출표 표시
                for sheet_name, data_list in st.session_state.extracted_data_by_sheet_page4.items():
                    st.subheader(f"■ {sheet_name} 추출값")
                    df = pd.DataFrame(data_list)
                    st.dataframe(df, use_container_width=True) # use_column_width 대신 use_container_width 사용

                # 날짜 입력 UI 준비
                # 추출된 항목 수에 맞춰 세션 상태의 date 값과 개별 위젯 키를 초기화합니다.
                for i in range(len(st.session_state.extracted_page4)):
                    if f"date1_page4_{i}" not in st.session_state: # key 변경
                        st.session_state[f"date1_page4_{i}"] = ""
                    if f"date2_page4_{i}" not in st.session_state: # key 변경
                        st.session_state[f"date2_page4_{i}"] = ""
                st.session_state.dates_page4 = [{"check": False, "date1": "", "date2": ""} # 세션 상태 키 변경
                                                for _ in st.session_state.extracted_page4] # 세션 상태 키 변경

    with left_col:
        # “최종 생성” 버튼 (추출 후에만 활성화)
        if st.session_state.extracted_page4: # 세션 상태 키 변경

            # --- 첫 번째 항목의 날짜 입력과 일괄 적용 체크박스 ---
            first_item = st.session_state.extracted_page4[0] # 세션 상태 키 변경
            c0, c1, c2, c3_for_checkbox = st.columns([2, 1, 1, 0.5])

            with c0:
                st.write(f"■ {first_item['사업명']}")
            with c1:
                st.text_input(
                    "점검일자",
                    value=st.session_state[f"date1_page4_{0}"], # 세션 상태에서 직접 값 가져옴
                    placeholder="YYYYMMDD 같은 형식으로 입력",
                    key=f"date1_page4_{0}" # key 변경
                )
            with c2:
                st.text_input(
                    "준공검사일",
                    value=st.session_state[f"date2_page4_{0}"], # 세션 상태에서 직접 값 가져옴
                    placeholder="YYYYMMDD 같은 형식으로 입력",
                    key=f"date2_page4_{0}" # key 변경
                )
            with c3_for_checkbox:
                st.markdown("<div style='margin-top: 30px;'></div>", unsafe_allow_html=True) # 정렬을 위한 마진
                st.checkbox(
                    "", # 레이블을 비워두고
                    key="apply_all_checkbox_page4", # key 변경
                    help="준공검사일과 점검일을 모든 항목에 일괄 적용합니다.",
                    label_visibility="hidden" # 레이블을 숨김
                )
                st.markdown("일괄<br>적용", unsafe_allow_html=True) # HTML 줄바꿈 태그 사용

            # --- "일괄 적용" 로직 (다른 위젯 생성 전에 실행) ---
            if st.session_state.apply_all_checkbox_page4: # key 변경
                first_d1 = st.session_state[f"date1_page4_{0}"]
                first_d2 = st.session_state[f"date2_page4_{0}"]
                for j in range(1, len(st.session_state.extracted_page4)): # 세션 상태 키 변경
                    # 실제로 값이 다를 때만 업데이트하여 불필요한 변경 방지
                    if st.session_state[f"date1_page4_{j}"] != first_d1: # key 변경
                        st.session_state[f"date1_page4_{j}"] = first_d1 # key 변경
                    if st.session_state[f"date2_page4_{j}"] != first_d2: # key 변경
                        st.session_state[f"date2_page4_{j}"] = first_d2 # key 변경

            # --- 두 번째 항목부터의 날짜 입력 ---
            for idx in range(1, len(st.session_state.extracted_page4)): # 세션 상태 키 변경
                item = st.session_state.extracted_page4[idx] # 세션 상태 키 변경
                # 모든 행에 동일한 4개의 컬럼을 생성하되, 마지막 컬럼(0.5 비율)은 비워둠
                c0, c1, c2, c3_spacer = st.columns([2, 1, 1, 0.5])

                with c0:
                    st.write(f"■ {item['사업명']}")
                with c1:
                    st.text_input(
                        "점검일자",
                        value=st.session_state[f"date1_page4_{idx}"], # key 변경
                        placeholder="YYYYMMDD 같은 형식으로 입력",
                        key=f"date1_page4_{idx}" # key 변경
                    )
                with c2:
                    st.text_input(
                        "준공검사일",
                        value=st.session_state[f"date2_page4_{idx}"], # key 변경
                        placeholder="YYYYMMDD 같은 형식으로 입력",
                        key=f"date2_page4_{idx}" # key 변경
                    )
                with c3_spacer:
                    pass # 이 컬럼은 공간 유지를 위해 비워둠

            # --- 최종적으로 HWP 생성에 사용될 st.session_state.dates 리스트에 값을 저장 ---
            # 모든 위젯 처리 및 일괄 적용 로직이 완료된 후에 실행
            for idx in range(len(st.session_state.extracted_page4)): # 세션 상태 키 변경
                st.session_state.dates_page4[idx]["date1"] = st.session_state[f"date1_page4_{idx}"] # 세션 상태 키 변경
                st.session_state.dates_page4[idx]["date2"] = st.session_state[f"date2_page4_{idx}"] # 세션 상태 키 변경

            # 실제 ZIP 생성
            if st.button("최종 생성", key="final_generate_btn_page4"): # key 변경
                buffer = io.BytesIO()

                def format_for_hwp(s: str) -> str:
                    if len(s) == 8 and s.isdigit():
                        y = s[:4]
                        m = str(int(s[4:6])) # 월에서 선행 0 제거
                        d = str(int(s[6:8])) # 일에서 선행 0 제거
                        return f"{y}. {m}. {d}." # 날짜 포맷 변경:YYYY. M. D.
                    return s

                with zipfile.ZipFile(buffer, "w") as zipf:
                    for idx_hwp, data in enumerate(st.session_state.extracted_page4): # Enumerate from 0
                        hwp = win32com.client.Dispatch("HWPFrame.HwpObject")
                        hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")
                        
                        # 현재 데이터의 source_sheet에 따라 템플릿 경로를 선택
                        current_template_path = template_paths_by_sheet.get(data["source_sheet"])
                        if not current_template_path or not os.path.exists(current_template_path):
                            st.error(f"템플릿 파일을 찾을 수 없습니다: {current_template_path} (시트: {data['source_sheet']}). 이 항목은 건너뜁니다.")
                            hwp.Quit() # HWP 객체 해제
                            continue # 다음 항목으로 넘어감

                        # hwp.Open 호출 방식 변경: 파일 형식("HWP")과 부가 옵션을 명시적으로 전달
                        hwp.Open(current_template_path, "HWP", "versionwarning:false")

                        # 현재 데이터의 source_sheet에 따라 compliance_groups를 선택
                        current_compliance_groups = compliance_groups_by_sheet.get(data["source_sheet"], full_compliance_groups)

                        # — 의무 태그 (AL, AM, AN 포함)
                        for j in range(1, len(current_compliance_groups)+1): # 선택된 그룹의 길이만큼 반복
                            tag = data.get(f"의무{j}")
                            if tag:
                                hwp.HAction.GetDefault("AllReplace", hwp.HParameterSet.HFindReplace.HSet)
                                hwp.HParameterSet.HFindReplace.FindString   = tag
                                hwp.HParameterSet.HFindReplace.ReplaceString = "○"
                                hwp.HParameterSet.HFindReplace.IgnoreMessage = 1
                                hwp.HAction.Execute("AllReplace", hwp.HParameterSet.HFindReplace.HSet)
                                # 같은 그룹 내 다른 태그들은 공백으로 처리
                                # 주의: current_compliance_groups[j-1]이 유효한지 확인
                                if j-1 < len(current_compliance_groups):
                                    for o in [t for t in current_compliance_groups[j-1][3] if t != tag]:
                                        hwp.HAction.GetDefault("AllReplace", hwp.HParameterSet.HFindReplace.HSet)
                                        hwp.HParameterSet.HFindReplace.FindString   = o
                                        hwp.HParameterSet.HFindReplace.ReplaceString = ""
                                        hwp.HParameterSet.HFindReplace.IgnoreMessage = 1
                                        hwp.HAction.Execute("AllReplace", hwp.HParameterSet.HFindReplace.HSet)

                        # — M1/M2 & 팀장님/과장님
                        for tag in ["[M1]","[M2]","[팀장님]","[과장님]"]:
                            rep = data.get(tag.strip("[]"), "")
                            hwp.HAction.GetDefault("AllReplace", hwp.HParameterSet.HFindReplace.HSet)
                            hwp.HParameterSet.HFindReplace.FindString   = tag
                            hwp.HParameterSet.HFindReplace.ReplaceString = str(rep)
                            hwp.HParameterSet.HFindReplace.IgnoreMessage = 1
                            hwp.HAction.Execute("AllReplace", hwp.HParameterSet.HFindReplace.HSet)

                        # [수정] — 사진우1~6을 사업명, 부서명 등으로 변경
                        pics = {
                            "[사업명]": data["사업명"],
                            "[부서명]": data["부서명"],
                            "[담당자]": data["담당자"],
                            "[착공일]": data["착공일"],
                            "[준공일]": data["준공일"],
                            "[사업비]": data["사업비"]
                        }
                        for tag, val in pics.items():
                            hwp.HAction.GetDefault("AllReplace", hwp.HParameterSet.HFindReplace.HSet)
                            hwp.HParameterSet.HFindReplace.FindString   = tag
                            hwp.HParameterSet.HFindReplace.ReplaceString = str(val) # 값을 문자열로 변환
                            hwp.HParameterSet.HFindReplace.IgnoreMessage = 1
                            hwp.HAction.Execute("AllReplace", hwp.HParameterSet.HFindReplace.HSet)

                        # — 점검일자/준공검사일
                        # idx_hwp는 0부터 시작하므로 그대로 인덱스로 사용
                        raw1 = st.session_state.dates_page4[idx_hwp]["date1"]
                        raw2 = st.session_state.dates_page4[idx_hwp]["date2"]
                        for tag, raw in {"[점검일자]": raw1, "[준공검사일]": raw2}.items():
                            hwp.HAction.GetDefault("AllReplace", hwp.HParameterSet.HFindReplace.HSet)
                            hwp.HParameterSet.HFindReplace.FindString   = tag
                            hwp.HParameterSet.HFindReplace.ReplaceString = format_for_hwp(raw)
                            hwp.HParameterSet.HFindReplace.IgnoreMessage = 1
                            hwp.HAction.Execute("AllReplace", hwp.HParameterSet.HFindReplace.HSet)

                        # 저장 & ZIP 추가
                        # 파일 이름을 사업명과 동일하게 변경합니다.
                        # 사업명에 특수문자가 포함될 경우를 대비하여 파일명으로 안전하게 사용할 수 있도록 처리합니다.
                        사업명_safe = "".join(c for c in data['사업명'] if c.isalnum() or c in (' ', '_', '-')).strip()
                        if not 사업명_safe: # 사업명이 비어있거나 파일명으로 부적합할 경우 대체 이름 사용
                            사업명_safe = f"점검표_파일_{idx_hwp + 1}"

                        # 파일명 앞에 [점검표X] 형식으로 시트 이름을 추가합니다.
                        output_filename = f"[{data['source_sheet']}]{사업명_safe}.hwp"
                        
                        # 임시 파일 경로를 생성 (파일을 직접 만들지는 않음)
                        tmp_dir = tempfile.gettempdir()
                        # 파일 이름 충돌을 피하기 위해 고유한 이름 생성
                        unique_filename = f"hwp_{os.urandom(8).hex()}.hwp"
                        tmp_path = os.path.join(tmp_dir, unique_filename)

                        # hwp.SaveAs 호출 방식 변경: 3개의 인자(경로, 포맷, 추가옵션)를 모두 전달
                        # 세 번째 인자로 빈 문자열("")을 전달하여 기본 옵션으로 저장
                        hwp.SaveAs(tmp_path, "HWP", "")
                        hwp.Quit()
                        
                        # 생성된 임시 파일을 ZIP에 추가
                        zipf.write(tmp_path, arcname=output_filename)

                        # 임시 파일 삭제
                        try:
                            os.remove(tmp_path)
                        except OSError as e:
                            st.warning(f"임시 파일 삭제 실패: {e}")

                buffer.seek(0)
                st.download_button(
                    "ZIP 다운로드 (날짜 포함)",
                    data=buffer,
                    file_name="output_with_dates.zip",
                    mime="application/zip",
                    key="download_zip_page4" # key 변경
                )

    with right_col:
        # 설명 이미지 제목을 제거하여 이미지가 더 위로 올라가도록 합니다.
        # st.markdown("### 설명 이미지") # 이 줄을 제거했습니다.
        # 이미지 경로를 사용자가 제공한 로컬 경로로 업데이트합니다.
        image_path = r"C:\Users\Owner\Desktop\사진우\AI\도급위탁용역자동화\image\image4\guide.png"
        if os.path.exists(image_path):
            st.image(image_path, use_container_width=True)
        else:
            st.error(f"설명 이미지를 찾을 수 없습니다: {image_path}")

    # COM 해제는 run() 함수 종료 시 수행
    # 스레드가 종료될 때 CoUninitialize가 호출되도록 보장
    try:
        pythoncom.CoUninitialize()
    except pythoncom.com_error as e:
        # 이미 해제된 경우 등 com_error가 발생할 수 있으나 무시해도 됨
        pass


# 이 부분이 추가되었습니다.
# page4.py 파일이 직접 실행될 때만 아래 코드가 실행됩니다.
if __name__ == "__main__":
    st.set_page_config(layout="wide", page_title="도급위탁용역 점검표 생성기 (시연용)")
    st.title("도급위탁용역 점검표 생성기 (단독 시연)")
    run() # run() 함수 호출
